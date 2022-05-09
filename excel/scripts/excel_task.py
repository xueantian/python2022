import openpyxl
from openpyxl.styles import Font,Alignment,Side,Border

# set the font
font=Font(name='Calibri',size=12,bold=True,color='ff0000')
font2=Font(name='Calibri',size=10,bold=True,color='ff0000')
alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
side=Side(style='thin',color='000000')
border=Border(left=side,right=side,top=side,bottom=side)

#open the same file
workbook=openpyxl.load_workbook('sample.xlsx')
sheet=workbook['Sheet2']

#get the value
A1_value=sheet['A1'].value

#print(A1_value)
rows=sheet[2]
lst=[]
for row in rows:
    lst.append(row.value)
#print(lst)

# write the lst to the new excel file

for i in range(3,16):

    write_workbook=openpyxl.Workbook()
    write_sheet=write_workbook.active

    #Merge and center A1
    write_sheet.merge_cells(start_row=1,end_row=1,start_column=1,end_column=20)
    write_sheet.row_dimensions[1].height=25.8
    write_sheet.row_dimensions[2].height=30.8

    write_sheet['A1']=A1_value
    write_sheet['A1'].font=font
    write_sheet['A1'].alignment=alignment

    #title
    write_sheet.append(lst)
    title_row=write_sheet[2]
    for row in title_row:

        row.font=font
        row.alignment=alignment
        row.border=border

    #set the width of column
    cols=write_sheet[2]
    for col in cols:
        write_sheet.column_dimensions[col.column_letter].width=10
    #print(write_sheet['A1'].value)
    write_sheet.column_dimensions['B'].width = 30
    write_sheet.column_dimensions['D'].width =30
    write_sheet.column_dimensions['E'].width = 20
    write_sheet.column_dimensions['S'].width = 50
    # input the value to excel
    write_sheet['A3'].value=1
    write_sheet['B3'].value=sheet['B'+str(i)].value
    write_sheet['C3'].value = sheet['C' + str(i)].value
    write_sheet['D3'].value = sheet['D' + str(i)].value
    write_sheet['E3'].value = sheet['E' + str(i)].value
    write_sheet['F3'].value = sheet['F' + str(i)].value
    write_sheet['G3'].value = sheet['G' + str(i)].value
    write_sheet['H3'].value = sheet['H' + str(i)].value
    write_sheet['I3'].value = sheet['I' + str(i)].value
    write_sheet['J3'].value = sheet['J' + str(i)].value
    write_sheet['K3'].value = sheet['K' + str(i)].value
    write_sheet['L3'].value = sheet['L' + str(i)].value
    write_sheet['M3'].value = sheet['M' + str(i)].value
    write_sheet['N3'].value = sheet['N' + str(i)].value
    write_sheet['O3'].value = sheet['O' + str(i)].value
    write_sheet['P3'].value = sheet['P' + str(i)].value
    write_sheet['Q3'].value = sheet['Q' + str(i)].value
    write_sheet['R3'].value = sheet['R' + str(i)].value
    write_sheet['S3'].value = sheet['S' + str(i)].value


    write_workbook.save(write_sheet['D3'].value+'.xlsx')


