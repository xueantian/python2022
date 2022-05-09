import openpyxl
from openpyxl.styles import Font,PatternFill
import re

workbook=openpyxl.load_workbook('ISR1K_1D_Scale_2_3.xlsx')

font=Font(color='ff0000')
patt=PatternFill(fill_type='solid',fgColor='fff000')

sheet=workbook['ISR1K - 1D Scale']

workbook.copy_worksheet(sheet)
sheet.title='new number'

for cols in workbook['new number'].columns:
    if re.search('.*4P.*',cols[0].value):
        for i in range(0,50):
            cols[i].fill=patt
            cols[i].font=font

workbook.save('new number.xlsx')



