import openpyxl
a=openpyxl.load_workbook('Cisco.xlsx')
sheet=a['Cisco']
print(sheet.dimensions)

# 按行row 读取
def read_rows():
    rows=sheet.iter_rows(min_row=2,max_row=10,min_col=1,max_col=8)
    for row in rows:
        for cell in row:
            print(cell.value)
        print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')

#按列读取
def read_cols():
    cols=sheet.iter_cols(min_row=2,max_row=10,min_col=1,max_col=8)
    for col in cols:
        for cell in col:
            print(cell.value)
        print('-----------------------')

if __name__ == '__main__':
   # read_cols()
    read_rows()
